"""Daily Progress Report: Layer 1 (auto-capture) + Layer 3 (export).

Layer 1 — record_change() is called from the SAME place a room/item slider
already gets set (progress.ProgressStore.set() / itemprog's per-room write).
It appends one dated fact: {date, service, floor, activity, room}. No new
screen, no typing — the engineer never sees this happen. Same (date, service,
floor, activity, room) logged twice in a day is a no-op (idempotent), so
dragging a slider back and forth doesn't duplicate a row.

This is deliberately NOT a snapshot-and-diff of cumulative % (that would need
re-deriving "what changed" from two full states, and silently misattributes a
correction to "new work"). It logs the actual write, at the moment it happens
-- the same signal the engineer already produces by dragging a slider, just
kept instead of overwritten.

Layer 3 — build_workbook() turns one day's (or a date range's) logged facts
into the exact Excel the site team already hand-builds: one colored section
per service, FLOOR merged down the left when an activity spans a floor,
ACTIVITY as free text, AS PER SCHEDULE / REMARKS left blank with a
Leading/On schedule/Lagging dropdown + auto color -- never invented, because
this system has no schedule baseline to compare against (see pnl.py's own
"never invent a number" rule -- same principle, applied here).

Nothing here touches the forecast engine or pnl.py. It only reads what
progress-setting already writes, and a project's structure tree for floor
ordering.
"""
import json
from datetime import date as _date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------
# Layer 1 — auto-capture. Pure data, no I/O here (the caller owns the file).
# --------------------------------------------------------------------------
def record_change(log, date_str, service, floor, activity, room, item=None, qty=None, unit=None):
    """Append (or UPDATE) one captured fact in `log` (a plain list, e.g.
    loaded from dpr_log.json). `item`/`qty`/`unit` are optional (older log
    entries on disk won't have them -- group_for_export() below tolerates
    that) -- when given, `item` is the BOQ item's own description and `qty`
    is the real per-room quantity this specific action represents (room_qty
    x the fraction just set -- see siteprogress.py's _log_dpr_change() for
    exactly how that's resolved; never a guess, and left None for the
    OVERALL/no-room case where no honest room-scoped quantity exists).

    Same (date, service, floor, activity, room, item) touched again the same
    day UPDATES that entry's qty in place rather than adding a second row --
    dragging one room's slider from 30% to 70% twice in a day must show 70%
    worth of qty once, not 30%+70% double-counted. Returns the list back (so
    a caller can `log = record_change(log, ...)` or ignore the return).
    """
    key = (date_str, service, floor, activity, room, item)
    for e in log:
        if (e["date"], e["service"], e["floor"], e["activity"],
            e.get("room"), e.get("item")) == key:
            e["qty"], e["unit"] = qty, unit
            return log
    log.append({"date": date_str, "service": service, "floor": floor,
               "activity": activity, "room": room, "item": item, "qty": qty, "unit": unit})
    return log


def entries_for_range(log, start_date, end_date=None):
    """{date_str: [entry, ...]} for every date in [start_date, end_date]
    (inclusive), in the order they were logged. end_date defaults to
    start_date (a single day)."""
    end_date = end_date or start_date
    out = {}
    for e in log:
        if start_date <= e["date"] <= end_date:
            out.setdefault(e["date"], []).append(e)
    return out


# --------------------------------------------------------------------------
# Grouping — the exact shape the sheet needs: per service, per floor, in the
# order floors first appear in the log (which is the order the engineer
# actually worked them, a perfectly reasonable default row order).
# --------------------------------------------------------------------------
def group_for_export(day_entries, all_services, leaf_label="ROOM"):
    """day_entries: the list for ONE date, from entries_for_range().
    all_services: every service this project tracks (so a service with zero
    changes still gets its own "NO ACTIVITY" section, matching the site
    team's own sheet, which lists HVAC even when nothing happened there).
    leaf_label: "ROOM" or "ZONE" (see dpr.LEAF_LABEL) -- the word used in
    the generated text ("...IN ROOM NO 5,6" vs "...IN ZONE NO 5,6"), picked
    from the project's structure.kind, never hardcoded.

    Groups by (floor, activity, item) now, not just (floor, activity) --
    two different BOQ items under the same activity (e.g. pipe + saddle
    both under "Wall Piping") must not have their room lists blended into
    one misleading combined list. Older log entries with no `item` (from
    before this field existed) fall back to one row per (floor, activity)
    exactly as before, item/qty columns simply blank.

    Returns {service: [(floor, activity_text, item, qty, unit), ...]}.
    `activity_text` is the activity name plus the rooms touched that day for
    that (floor, activity, item), comma-joined -- built from real logged
    rooms, never guessed. `qty` is the SUM of each captured action's own qty
    for that (floor, activity, item) group -- see record_change()'s own note
    on why repeat touches of the SAME room update in place rather than
    summing (so this sum is real, not double-counted); different rooms
    genuinely do add up.
    """
    by_service = {s: {} for s in all_services}
    order = {s: [] for s in all_services}
    for e in day_entries:
        svc = e["service"]
        if svc not in by_service:
            by_service[svc] = {}
            order[svc] = []
        key = (e["floor"], e["activity"], e.get("item"))
        if key not in by_service[svc]:
            by_service[svc][key] = {"rooms": [], "qty": 0.0, "unit": e.get("unit"), "has_qty": False}
        slot = by_service[svc][key]
        if key not in order[svc]:
            order[svc].append(key)
        if e["room"] and e["room"] not in slot["rooms"]:
            slot["rooms"].append(e["room"])
        if e.get("qty") is not None:
            slot["qty"] += e["qty"]
            slot["has_qty"] = True
            slot["unit"] = e.get("unit") or slot["unit"]

    out = {}
    for svc in all_services:
        rows = []
        for (floor, activity, item) in order.get(svc, []):
            slot = by_service[svc][(floor, activity, item)]
            rooms = slot["rooms"]
            text = (f"{activity.upper()} IN {leaf_label} NO {','.join(rooms)}"
                   if rooms else activity.upper())
            qty = round(slot["qty"], 3) if slot["has_qty"] else None
            rows.append((floor, text, item, qty, slot["unit"]))
        out[svc] = rows
    return out


# --------------------------------------------------------------------------
# Layer 3 — the formatter. Pure function: grouped data in, Workbook out.
# --------------------------------------------------------------------------
FONT = "Arial"
# Matches siteprogress.js's own leafLabel(kind) convention (Room vs Zone) --
# same idea, one level up the tree: the label for a room's immediate/joined
# container path, picked from structure.kind, never hardcoded to "FLOOR".
LOCATION_LABEL = {"hotel": "FLOOR", "mall": "LEVEL", "hospital": "WING / FLOOR", "custom": "LOCATION"}
# Matches siteprogress.js's leafLabel(kind): "Zone" only for a mall, "Room"
# for hotel/hospital/custom -- used inside the generated activity text
# ("...IN ROOM NO 5,6" / "...IN ZONE NO 5,6"), not the column header.
LEAF_LABEL = {"hotel": "ROOM", "mall": "ZONE", "hospital": "ROOM", "custom": "ROOM"}
_thin = Side(style="thin", color="000000")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
SECTION_FILL = {  # cycle through these for services beyond the first 4, so a
                  # 5th service never crashes -- just repeats the palette.
    0: "BDD7EE", 1: "FBE0CE", 2: "FFE699", 3: "C6E0B4", 4: "D9D2E9",
}


def _write_day_block(ws, row, date_str, grouped, location_label="FLOOR"):
    ws.cell(row, 1, "DATE:").font = Font(name=FONT, bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row, 2, date_str).font = Font(name=FONT, size=11)
    for c in range(1, 6):
        ws.cell(row, c).border = BORDER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    t = ws.cell(row, 1, "DAILY WORK UPDATES")
    t.font = Font(name=FONT, bold=True, size=11)
    t.alignment = Alignment(horizontal="center")
    for c in range(1, 6):
        ws.cell(row, c).border = BORDER
    row += 1

    for i, (service, items) in enumerate(grouped.items()):
        fill = PatternFill("solid", fgColor=SECTION_FILL[i % len(SECTION_FILL)])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        s = ws.cell(row, 1, service.upper())
        s.font = Font(name=FONT, bold=True, size=11)
        s.alignment = Alignment(horizontal="center")
        for c in range(1, 6):
            ws.cell(row, c).fill = fill
            ws.cell(row, c).border = BORDER
        row += 1

        headers = [location_label, "ACTIVITY", "ITEM", "QTY EXECUTED", "REMARKS"]
        for c, h in zip(range(1, 6), headers):
            hc = ws.cell(row, c, h)
            hc.font = Font(name=FONT, bold=True, size=10)
            hc.alignment = Alignment(horizontal="center")
            hc.border = BORDER
        row += 1

        if not items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            na = ws.cell(row, 1, "NO ACTIVITY")
            na.font = Font(name=FONT, size=10)
            na.alignment = Alignment(horizontal="center")
            for c in range(1, 6):
                ws.cell(row, c).border = BORDER
            row += 1
            continue

        floor_start_row = row
        prev_floor = None
        for floor, activity_text, item, qty, unit in items:
            if floor != prev_floor and prev_floor is not None and row - 1 > floor_start_row:
                ws.merge_cells(start_row=floor_start_row, start_column=1, end_row=row - 1, end_column=1)
            if floor != prev_floor:
                floor_start_row = row
            fc = ws.cell(row, 1, floor)
            fc.font = Font(name=FONT, size=10)
            fc.alignment = Alignment(horizontal="center", vertical="center")
            ac = ws.cell(row, 2, activity_text)
            ac.font = Font(name=FONT, size=10)
            ic = ws.cell(row, 3, item or "—")
            ic.font = Font(name=FONT, size=10, italic=(not item), color="9A9890" if not item else "000000")
            if qty is not None:
                qc = ws.cell(row, 4, f"{qty:,.2f} {unit}" if unit else round(qty, 2))
                qc.alignment = Alignment(horizontal="right")
                qc.font = Font(name=FONT, size=10)
            else:
                qc = ws.cell(row, 4, "—")
                qc.font = Font(name=FONT, size=10, italic=True, color="9A9890")
            for c in range(1, 6):
                ws.cell(row, c).border = BORDER
            prev_floor = floor
            row += 1
        if row - 1 > floor_start_row:
            ws.merge_cells(start_row=floor_start_row, start_column=1, end_row=row - 1, end_column=1)
    return row


def build_workbook(days, location_label="FLOOR"):
    """days: [(date_str, grouped), ...] -- grouped is group_for_export()'s
    output for that date, in the order dates should stack. One sheet, one
    block per day, exactly matching the site team's own printed layout.

    location_label: the header text for the first column. record_change()/
    group_for_export() already treat "floor" as an opaque grouping string --
    they work unchanged for a mall (Level > Zone) or a hospital (Wing > Floor
    > Room, one level deeper than hotel/mall) as long as the CALLER passes
    the room's full container path joined into one string (e.g. "Wing A ·
    Floor 3" for a hospital room, from Structure.rooms()'s own `path`, minus
    the project name at path[0]) -- this label is only the column header text,
    picked from the project's structure kind (LOCATION_LABEL below), the same
    way siteprogress.js already picks "Room" vs "Zone" from structure.kind."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DPR"
    ws.sheet_view.showGridLines = False
    # column 1's width must fit the LONGEST location string that will
    # actually appear -- "13TH" (hotel) and "Wing A · Floor 3" (hospital)
    # need very different widths; sizing to a fixed 9 clipped the hospital
    # case. Widened to whichever is longer: the header label or any day's
    # actual location values.
    longest = len(location_label)
    for _, grouped in days:
        for rows in grouped.values():
            for floor, *_rest in rows:
                longest = max(longest, len(floor))
    widths = {1: max(9, longest + 2), 2: 40, 3: 26, 4: 16, 5: 30}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 1
    for date_str, grouped in days:
        row = _write_day_block(ws, row, date_str, grouped, location_label)
        row += 1   # blank row between day blocks
    return wb


# --------------------------------------------------------------------------
# Summary + item-detail sheets -- an MD opens the file and the FIRST tab
# already answers "where do we stand", using the exact same numbers the
# Overall dashboard already shows (pnl.rollup_pnl / project_pnl / itemprog's
# room_buckets). Nothing new is computed here -- this only formats numbers
# those already-tested functions produce. The site team's own existing
# export template (Services/Item/Planned/Used so far/Remaining/Done/In
# progress/Pending) becomes the "Item detail" tab, finally populated with
# real numbers instead of shipping as a blank header.
#
# Room-count totals in Summary are real SUMIF formulas over the Item detail
# sheet (both sheets share that column), not re-passed numbers -- open the
# cell and Excel shows exactly where it came from. ₹ figures stay computed
# values: Item detail's own template (the site team's blank original) never
# had a rate/value column, so there is nothing on that sheet for a ₹ formula
# to reference -- inventing one just to force a formula would be a bigger
# change to their template than asked. A note on the sheet says so plainly,
# per the "document every hardcoded number" rule.
# --------------------------------------------------------------------------
HEADER_FILL = "2C3E50"
HEADER_FONT_COLOR = "FFFFFF"
ZEBRA_FILL = "F2F2F2"
HERO_FILL = "EEF1F4"
_thin_gray = Side(style="thin", color="B0B0B0")
GRID_BORDER = Border(left=_thin_gray, right=_thin_gray, top=_thin_gray, bottom=_thin_gray)


def _header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row, col_start + i, h)
        c.font = Font(name=FONT, bold=True, size=10, color=HEADER_FONT_COLOR)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = GRID_BORDER


def _zebra(ws, row, n_cols, col_start=1):
    if row % 2 == 0:
        for c in range(col_start, col_start + n_cols):
            cell = ws.cell(row, c)
            if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor=ZEBRA_FILL)


def build_summary_sheet(wb, date_label, project_totals, by_service, generated_at=None):
    """project_totals: pnl.project_pnl()'s dict (done_value, remaining_value,
    full_value, waste_value, pct_value_done, ...).
    by_service: {service: {done_value, remaining_value}} -- ₹ only; room
    counts are pulled from the Item detail sheet via formula, not passed in
    here, so the two sheets can never quietly disagree.
    generated_at: 'YYYY-MM-DD HH:MM' string -- when this file was actually
    produced, so a downloaded report is never mistaken for a live view."""
    ws = wb.create_sheet("Summary", 0)   # first tab -- the thing an MD sees on open
    ws.sheet_view.showGridLines = False
    for c, w in {1: 16, 2: 16, 3: 16, 4: 16, 5: 12, 6: 12, 7: 12}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells("A1:G1")
    h = ws.cell(1, 1, f"Site progress summary — {date_label}")
    h.font = Font(name=FONT, bold=True, size=14, color=HEADER_FILL)
    ws.row_dimensions[1].height = 22
    if generated_at:
        ws.merge_cells("A2:G2")
        g = ws.cell(2, 1, f"Generated {generated_at}")
        g.font = Font(name=FONT, size=9, italic=True, color="9A9890")

    hero = [("Value complete", project_totals.get("pct_value_done", 0) / 100, "0.0%"),
           ("Done", project_totals.get("done_value"), None),
           ("Remaining", project_totals.get("remaining_value"), None),
           ("Material waste", project_totals.get("waste_value"), None)]
    for i, (label, val, fmt) in enumerate(hero):
        c0 = 1 + i * 2
        for r in (3, 4):
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 1)
            cell = ws.cell(r, c0)
            cell.fill = PatternFill("solid", fgColor=HERO_FILL)
            ws.cell(r, c0 + 1).fill = PatternFill("solid", fgColor=HERO_FILL)
            if i < len(hero) - 1:   # right divider so each card reads as its own tile
                ws.cell(r, c0 + 1).border = Border(right=Side(style="thin", color="C9CDD3"))
        lc = ws.cell(3, c0, label)
        lc.font = Font(name=FONT, size=10, color="5F5E5A")
        lc.alignment = Alignment(horizontal="left", indent=1)
        vc = ws.cell(4, c0, val if fmt else round(val or 0))
        vc.font = Font(name=FONT, bold=True, size=16, color=HEADER_FILL)
        vc.number_format = fmt or '"₹"#,##0'
        vc.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 24

    row = 6
    headers = ["Service", "% complete", "Done (₹)", "Remaining (₹)", "Rooms done", "In progress", "Pending"]
    _header_row(ws, row, headers)
    ws.row_dimensions[row].height = 18
    row += 1
    first_data_row = row
    for svc, t in by_service.items():
        done_v = t.get("done_value") or 0
        rem_v = t.get("remaining_value") or 0
        planned_v = done_v + rem_v
        ws.cell(row, 1, svc).border = GRID_BORDER
        pc = ws.cell(row, 2, f"=IFERROR(C{row}/(C{row}+D{row}),0)")
        pc.number_format = "0.0%"
        pc.border = GRID_BORDER
        dv = ws.cell(row, 3, round(done_v))
        dv.number_format = '"₹"#,##0'
        dv.border = GRID_BORDER
        rv = ws.cell(row, 4, round(rem_v))
        rv.number_format = '"₹"#,##0'
        rv.border = GRID_BORDER
        # real formulas -- summed straight off the Item detail sheet's own
        # Done/In progress/Pending columns for this service, so Summary can
        # never quietly drift from the sheet it is summarising.
        for c, col_letter in ((5, "F"), (6, "G"), (7, "H")):
            f = ws.cell(row, c, f"=SUMIF('Item detail'!$A:$A,A{row},'Item detail'!${col_letter}:${col_letter})")
            f.number_format = "#,##0"
            f.border = GRID_BORDER
        for c in range(1, 8):
            ws.cell(row, c).font = Font(name=FONT, size=10)
        _zebra(ws, row, 7)
        row += 1
    ws.conditional_formatting.add(f"B{first_data_row}:B{row - 1}",
        CellIsRule(operator="lessThan", formula=["0.3"], font=Font(name=FONT, size=10, color="A32D2D")))
    ws.conditional_formatting.add(f"B{first_data_row}:B{row - 1}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.7"], font=Font(name=FONT, size=10, color="3B6D11")))
    ws.freeze_panes = f"A{first_data_row}"

    note = ws.cell(row + 1, 1,
        "₹ figures: computed at export time from each item's BOQ rate x installation % (see the Payment terms "
        "setting). Room counts: live SUMIF formulas off the Item detail sheet, columns F–H.")
    note.font = Font(name=FONT, size=9, italic=True, color="888780")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 2, end_column=7)
    ws.row_dimensions[row + 1].height = 28
    return ws


def _inr(v):
    if v is None:
        return "—"
    v = float(v)
    if abs(v) >= 1e7:
        return f"₹{v/1e7:.2f} Cr"
    if abs(v) >= 1e5:
        return f"₹{v/1e5:.2f} L"
    return f"₹{v:,.0f}"


def build_item_detail_sheet(wb, item_rows):
    """item_rows: [{service, item, planned, used, remaining, done,
    in_progress, pending}, ...] -- same 8 columns as the site team's own
    Report_Format.xlsx, done/in_progress/pending are real room counts from
    itemprog.room_buckets (the caller's job to supply -- this only formats),
    never left blank once real progress data exists for that item."""
    ws = wb.create_sheet("Item detail")
    ws.sheet_view.showGridLines = False
    headers = ["Services", "Item", "Planned", "Used so far", "Remaining", "Done", "In progress", "Pending"]
    widths = [14, 36, 10, 12, 10, 8, 10, 8]
    _header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 18
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    numeric_cols = {3, 4, 5, 6, 7, 8}
    for r, item in enumerate(item_rows, start=2):
        vals = [item.get(k) for k in ("service", "item", "planned", "used", "remaining",
                                      "done", "in_progress", "pending")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10)
            cell.border = GRID_BORDER
            if c in numeric_cols:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        _zebra(ws, r, 8)
    last_row = len(item_rows) + 1
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:H{last_row}"
    ws.freeze_panes = "A2"
    return ws


# --------------------------------------------------------------------------
# Activity completion -- mirrors the site team's own real "Activity
# Completion Summary — Auto from Room Detail" report exactly: per service, a
# dark section header, then one row per activity (Total Rooms / Done /
# Partial / Pending / %Done / %Pending), a bold TOTAL row, and %Done shaded
# green/amber/red by threshold -- same color logic real construction status
# reports use. Floor-level rows sit UNDER each activity, collapsed by
# default via Excel's own outline grouping (click the [+] to expand) so the
# sheet reads clean at a glance but the "13th floor: 106 of 108 done" detail
# is one click away, never deleted.
#
# Source: progress.activity_completion() off the real Room-Detail tick
# grid (progress.parquet) -- a DIFFERENT signal from the item-level BOQ
# sliders that drive the Summary/Item detail/Daily updates sheets. Both are
# real, both are already-tracked data; this sheet is the one that reads off
# the tick grid specifically, because that is what a room-completion
# percentage actually means on this site.
# --------------------------------------------------------------------------
_PCT_GREEN, _PCT_AMBER, _PCT_RED = "C6E0B4", "FFE699", "F4C7C3"


def _pct_fill(pct):
    if pct >= 90:
        return _PCT_GREEN
    if pct >= 70:
        return _PCT_AMBER
    return _PCT_RED


def build_activity_completion_sheet(wb, completion, completion_by_floor=None):
    """completion: progress.activity_completion()'s output (by_floor=False).
    completion_by_floor: the same call with by_floor=True, optional -- when
    given, each activity gets its per-floor rows nested (collapsed) beneath
    it. Both come from the SAME progress.parquet read, so they can never
    disagree with each other."""
    ws = wb.create_sheet("Activity completion")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False   # floor rows collapse UNDER their activity row
    widths = {1: 34, 2: 12, 3: 10, 4: 10, 5: 10, 6: 10, 7: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 1
    for i, (service, activities) in enumerate(completion.items()):
        fill = PatternFill("solid", fgColor=SECTION_FILL[i % len(SECTION_FILL)])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        s = ws.cell(row, 1, f"{service.upper()} — ACTIVITY COMPLETION (ALL FLOORS COMBINED)")
        s.font = Font(name=FONT, bold=True, size=11)
        s.alignment = Alignment(horizontal="center")
        for c in range(1, 8):
            ws.cell(row, c).fill = fill
            ws.cell(row, c).border = GRID_BORDER
        row += 1

        headers = ["Activity", "Total rooms", "✓ Done", "~ Partial", "✗ Pending", "% Done", "% Pending"]
        _header_row(ws, row, headers)
        row += 1

        tot = {"total": 0, "done": 0, "partial": 0, "pending": 0}
        by_floor_this_svc = (completion_by_floor or {}).get(service, {})
        for act, m in activities.items():
            vals = [act, m["total"], m["done"], m["partial"], m["pending"],
                   m["pct_done"] / 100, m["pct_pending"] / 100]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row, c, v)
                cell.font = Font(name=FONT, size=10)
                cell.border = GRID_BORDER
                if c in (2, 3, 4, 5):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                if c in (6, 7):
                    cell.number_format = "0.00%"
            ws.cell(row, 6).fill = PatternFill("solid", fgColor=_pct_fill(m["pct_done"]))
            for k in tot:
                tot[k] += m[k]
            activity_row = row
            row += 1

            floors = by_floor_this_svc.get(act)
            if floors:
                for floor, fm in floors.items():
                    frow = [f"    {floor}", fm["total"], fm["done"], fm["partial"], fm["pending"],
                           fm["pct_done"] / 100, fm["pct_pending"] / 100]
                    for c, v in enumerate(frow, start=1):
                        cell = ws.cell(row, c, v)
                        cell.font = Font(name=FONT, size=9, color="6B6A66")
                        cell.border = GRID_BORDER
                        if c in (2, 3, 4, 5):
                            cell.alignment = Alignment(horizontal="right")
                        if c in (6, 7):
                            cell.number_format = "0.00%"
                    ws.row_dimensions[row].outlineLevel = 1
                    ws.row_dimensions[row].hidden = True
                    row += 1

        total_pct_done = round(100.0 * tot["done"] / tot["total"], 2) if tot["total"] else 0.0
        total_pct_pending = round(100.0 * tot["pending"] / tot["total"], 2) if tot["total"] else 0.0
        trow = [f"{service.upper()} TOTAL", tot["total"], tot["done"], tot["partial"], tot["pending"],
               total_pct_done / 100, total_pct_pending / 100]
        for c, v in enumerate(trow, start=1):
            cell = ws.cell(row, c, v)
            cell.font = Font(name=FONT, bold=True, size=10, color=HEADER_FILL)
            cell.border = GRID_BORDER
            if c in (2, 3, 4, 5):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            if c in (6, 7):
                cell.number_format = "0.00%"
        ws.cell(row, 6).fill = PatternFill("solid", fgColor=_pct_fill(total_pct_done))
        row += 2   # blank separator row before the next service
    ws.freeze_panes = "A1"
    return ws


def build_full_export(date_label, project_totals, by_service, item_rows, days, location_label="FLOOR",
                      completion=None, completion_by_floor=None, generated_at=None):
    """The complete DPR: Summary (first, MD-facing) + Activity completion
    (real Room-Detail tick data, when supplied) + Item detail (the site
    team's own existing template, mapped items only) + the day-by-day
    narrative log."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_summary_sheet(wb, date_label, project_totals, by_service, generated_at=generated_at)
    if completion:
        build_activity_completion_sheet(wb, completion, completion_by_floor=completion_by_floor)
    build_item_detail_sheet(wb, item_rows)
    narrative = build_workbook(days, location_label=location_label)
    src = narrative.active
    dst = wb.create_sheet("Daily updates")
    import copy
    for row in src.iter_rows():
        for cell in row:
            nc = dst.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                nc.font = copy.copy(cell.font)
                nc.border = copy.copy(cell.border)
                nc.fill = copy.copy(cell.fill)
                nc.alignment = copy.copy(cell.alignment)
    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col].width = dim.width
    for mc in src.merged_cells.ranges:
        dst.merge_cells(str(mc))
    dst.sheet_view.showGridLines = False
    return wb
